import openpyxl

book = openpyxl.load_workbook("//home//liby//Documents//exceldemo.xlsx")#go to the excel sheet
Dict = {}
sheet = book.active#to get the control on the active sheet in excel
cell = sheet.cell(row=1,column=3)#cell starts with row and column as 1
print(cell.value)#to print that cell value
sheet.cell(row=1,column=5).value = 'Rahul'#to write a value into the excel
print(sheet.cell(row=1,column=5).value)
print(sheet.max_row)#to print max-row
print(sheet.max_column)#to print max-column
print(sheet['A6'].value)#another way to get the cell value
############to print all the values in the excel
"""
for i in range( 1,sheet.max_row+1):#i will iterate from 1st cell to last row cell
    for j in range(1,sheet.max_column+1):#j will iterate from 1st cell to last column cell
        print(sheet.cell(row=i,column=j).value)
"""
############to print a particular row in excel
for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == 'Testcase2':
        for j in range(2,sheet.max_column+1):
            #print(sheet.cell(row=i,column=j).value)#printing all values from the specific row and column
            ###########creating excel data into a dictionary
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
print(Dict)