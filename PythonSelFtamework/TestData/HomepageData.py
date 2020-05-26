import openpyxl


class Testhomepagedata:
    test_homedata = [{"Name": "Liby", "Email": "libyelsa795@gmail.com", "Password": "mary@12", "Gender": "Female"}, {"Name": "Rahul","Email": "rahul795@gmail.com","Password": "mary@12","Gender": "Male"}]

@staticmethod
def get_data(testcase_value):
    Dict={}
    book = openpyxl.load_workbook("//home//liby//Documents//exceldemo.xlsx")
    sheet = book.active
    for i in range(1, sheet.max_row + 1):
        if sheet.cell(row=i, column=1).value == testcase_value:
            for j in range(2, sheet.max_column + 1):
                # print(sheet.cell(row=i,column=j).value)#printing all values from the specific row and column
                ###########creating excel data into a dictionary
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    return [Dict]#to return as list